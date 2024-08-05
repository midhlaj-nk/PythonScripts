import psycopg2
from openpyxl import Workbook


def export_tables_to_excel(odoo_version, tables, issues=None):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Tables in Odoo {odoo_version}"

    for index, table in enumerate(tables, start=1):
        ws.cell(row=index, column=1, value=table[0])

    if issues:
        ws_issues = wb.create_sheet(title=f"Issues in Odoo {odoo_version}")
        for index, issue in enumerate(issues, start=1):
            ws_issues.cell(row=index, column=1, value=issue)

    wb.save(f"odoo_{odoo_version}_tables.xlsx")
    print(f"Tables in Odoo {odoo_version} exported to odoo_{odoo_version}_tables.xlsx")


def copy_tables_data(odoo_16_tables, odoo_17_tables):
    issues = []

    for table_16 in odoo_16_tables:
        if table_16 in odoo_17_tables:
            try:
                # Establish connections for both databases
                conn_16 = psycopg2.connect(
                    dbname=odoo_16_db_name,
                    user=odoo_16_db_user,
                    password=odoo_16_db_password,
                    host='localhost',
                    port='5432'
                )
                conn_17 = psycopg2.connect(
                    dbname=odoo_17_db_name,
                    user=odoo_17_db_user,
                    password=odoo_17_db_password,
                    host='localhost',
                    port='5432'
                )

                cursor_16 = conn_16.cursor()
                cursor_17 = conn_17.cursor()

                # Fetch data from Odoo 16 table
                cursor_16.execute(f"SELECT * FROM {table_16};")
                data_16 = cursor_16.fetchall()

                # Fetch column names from both tables
                cursor_16.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_16}';")
                columns_16 = [col[0] for col in cursor_16.fetchall()]

                cursor_17.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_16}';")
                columns_17 = [col[0] for col in cursor_17.fetchall()]

                # Construct the INSERT INTO query for Odoo 17
                insert_query = f"INSERT INTO {table_16} ({', '.join(columns_17)}) VALUES "

                # Generate values placeholders for the query
                values_placeholder = ", ".join(["%s"] * len(columns_17))

                # Construct the full INSERT INTO query
                insert_query += f"({values_placeholder});"

                # Execute the INSERT INTO query for each row
                for row in data_16:
                    cursor_17.execute(insert_query, row)

                # Commit the changes
                conn_17.commit()

                # Close connections
                cursor_16.close()
                cursor_17.close()
                conn_16.close()
                conn_17.close()

                print(f"Data copied successfully from {table_16} in Odoo 16 to Odoo 17")
            except psycopg2.Error as e:
                issues.append(f"Error copying data from {table_16}: {e}")
        else:
            issues.append(f"{table_16} doesn't exist in Odoo 17")

    return issues



def list_tables(odoo_version, db_name, db_user, db_password, db_host='localhost', db_port='5432'):
    try:
        # Establish connection
        conn = psycopg2.connect(
            dbname=db_name,
            user=db_user,
            password=db_password,
            host=db_host,
            port=db_port
        )
        cursor = conn.cursor()

        # List tables
        cursor.execute(
            "SELECT table_name FROM information_schema.tables WHERE table_schema = 'public' ORDER BY table_name;")
        tables = cursor.fetchall()

        # Close connections
        cursor.close()
        conn.close()

        return tables

    except psycopg2.Error as e:
        print("Error: Unable to connect to the database.")
        print(e)
        return []


# Odoo 16 database details
odoo_16_db_name = "odoo_16_data"
odoo_16_db_user = "odoo16"
odoo_16_db_password = "cool"

# Odoo 17 database details
odoo_17_db_name = "odoo_17_community"
odoo_17_db_user = "odoo17"
odoo_17_db_password = "cool"

# Get tables for Odoo 16
odoo_16_tables = list_tables("16", odoo_16_db_name, odoo_16_db_user, odoo_16_db_password)

# Get tables for Odoo 17
odoo_17_tables = list_tables("17", odoo_17_db_name, odoo_17_db_user, odoo_17_db_password)

# Copy tables data from Odoo 16 to Odoo 17 and identify issues
copy_issues = copy_tables_data(odoo_16_tables, odoo_17_tables)

# Export tables and issues to Excel
export_tables_to_excel("16", odoo_16_tables)
export_tables_to_excel("17", odoo_17_tables, copy_issues)
